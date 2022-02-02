from argparse import ArgumentParser
import re
from typing import Any
from pathlib import Path
import json

import openpyxl
import openpyxl.styles as styles
import openpyxl.worksheet.worksheet as worksheet
import openpyxl.cell.cell as cell
import dictknife 
import yaml

START_ROW = 3

def generate_testlist(lines: str, base: str=".") -> list[dict[list[str] | dict[str]]]:
  """
  Markdownデータより、試験項目用リストを作成する。
  
  Parameters
  ----
  lines: 試験項目データを含むMarkdownデータ
  basedir: プリプロセッサ実行時の基準ディレクトリパス

  Returns
  ----
  試験項目を含む構造体
  """
  level = 0
  examsmap = []
  itemmap = []
  previoustest = {}
  currenttest = {}
  section = ""
  textbuf = []
  basedir = Path(base)
  text = ""
  for line in lines.split("\n"):
    if m := re.match(r"\s+&(\w+)\((.*?)\)$", line):
      # preprocessor
      argument = json.loads(m[2])
      match m[1]:
        case "include":
          with open(basedir / argument["name"], mode="r", encoding="utf-8") as f:
            s = f.read()
            for n, v in argument.items():
              s = s.replace(f"//**{n}**//", v)
            text += s
    else:
      text += f"{line}\n"
  lines = text
  for line in lines.split("\n"):
    if m := re.match(r"^\s*(#+)\s*(.*)$", line):
      # change item
      if textbuf != [] and section != "":
        currenttest[section] = textbuf
      if itemmap != [] and currenttest != {}:
        examsmap.append({
          "items": itemmap.copy(),
          "exams": currenttest,
        })
      # new item
      ml = len(m[1])
      if level == ml:
        itemmap[-1] = m[2]
      elif level + 1 == ml:
        itemmap.append(m[2])
        level+=1
      elif level > ml:
        itemmap = itemmap[0:ml]
        itemmap[-1] = m[2]
        level=ml
      else:
        raise Exception("Incorrect test vote data.")
      section = ""
      if currenttest != {}:
        previoustest = currenttest.copy()
      currenttest = {}
      textbuf = []
    elif m := re.match(r"^\s*::\s*(.*?)\s*(&&)?$", line):
      # change section
      if textbuf != [] and section != "":
        currenttest[section] = textbuf
      # new section
      section = m[1]
      if m[2] == "&&" and section in previoustest:
        textbuf = previoustest[section]
      else:
        textbuf = []
    elif re.match(r"^\s*$", line):
      # ignore blank line.
      pass
    else:
      textbuf += [line.strip()]
  if textbuf != [] and section != "":
    currenttest[section] = textbuf
  if itemmap != [] and currenttest != {}:
    examsmap.append({
      "items": itemmap.copy(),
      "exams": currenttest,
    })
  return examsmap

def cells_normalization(testitemslabel: list[str], examsmap: list[dict[list[str] | dict[str]]]) -> list[list[str]]:
  """
  試験データの正規化を行う
  
  Parameters
  ----
  testitemslabel: 試験項目タイトルを示すラベル
  examsmap: generate_testlistメソッドの出力値

  Returns
  ----
  試験項目を示すテーブルデータ。
  """
  tilcount = len(testitemslabel)
  lines = []
  header = {}
  ids = [0] * tilcount
  prevrowname = [""] * tilcount
  for exam in examsmap:
    line = []
    # itemname
    if tilcount > len(exam["items"]):
      line = exam["items"] + [""] * (tilcount - len(exam["items"]))
    elif tilcount < len(exam["items"]):
      raise Exception("Incorrect test data.") 
    else:
      line = exam["items"]
    # name count
    changed = False
    for i, n in enumerate(exam["items"]):
      if prevrowname[i] != n and not changed:
        ids[i] += 1
        if i < tilcount:
          ids[(i + 1):] = [1] * (tilcount - i - 1)
        changed = True
      prevrowname[i] = n
    line.insert(0, "-".join(map(str, ids)))
    # preload exams
    for ex in exam["exams"].items():
      if not ex[0] in header:
        header[ex[0]] = len(header)
    examdata = [""] * len(header)
    for n, v in exam["exams"].items():
      examdata[header[n]] = v
    lines.append(line + examdata)

  line = ["No"] + testitemslabel
  for h in header.keys():
    line.append(h)
  lines.insert(0, line)

  # Leveling of the number of arrays
  l = len(max(lines, key=len))
  for i, line in enumerate(lines):
    if len(line) < l: lines[i] += [""] * (l - len(line))
  return lines
  
def add_examcells(examinfo: dict[str | int | list[str]], cells: list[list[str]]):
  """
  試験実施確認用セルを作成する

  Parameters
  ----
  examinfo: 試験実施確認用のデータを示す構造体
    PrintCount: Excel表に出力する試験実施の試行回数。試行回数分の列が追加される
    Labels: 試験実施のラベル(配列)
  cells: cells_normalizationの出力値

  Returns
  ----
  試験項目を示すテーブルデータ。
  """
  ii = len(max(cells, key=len))
  for c in range(examinfo["PrintCount"]):
    for l in examinfo["Labels"]:
      for i, line in enumerate(cells):
        line.insert(ii, l if i == 0 else "")
        ii += 1
  return cells
  
def rearrange_cells(headers: dict[Any], cells: list[list[str]], arrangeitems: list[str]) -> list[list[str]]:
  """
  テーブルを並び替える。なお、このメソッドは二回以上呼び出しできない。

  Parameters
  ----
  headers: 設定構造体
  cells: 試験項目を示すテーブルデータ。
  arrangeitems: 並び順を示すリスト。以下の文字列が必ず含まれる必要がある
    no: 試験番号列
    itemname: 試験項目名列
    content: 試験内容列
    results: 試験結果列

  Returns
  ----
  試験項目を示すテーブルデータ。
  """
  no  = cells[0].index("No")
  ins = cells[0].index(headers["TestItemsLabel"][0])
  ine = ins + len(headers["TestItemsLabel"]) - 1
  cont= ine + 1
  ress= cells[0].index(headers["TestResult"]["Labels"][0])
  rese= ress + len(headers["TestResult"]["Labels"]) * headers["TestResult"]["PrintCount"]
  ano = []
  ain = []
  acon= []
  ares= []
  for c in cells:
    ano.append(c[no:no + 1])
    ain.append(c[ins:ine + 1])
    acon.append(c[cont:ress])
    ares.append(c[ress:rese])
  result = [[] for i in range(len(cells))]
  for n in arrangeitems:
    match n:
      case "no":
        for i, item in enumerate(ano):
          result[i] += item
      case "itemname":
        for i, item in enumerate(ain):
          result[i] += item
      case "content":
        for i, item in enumerate(acon):
          result[i] += item
      case "results":
        for i, item in enumerate(ares):
          result[i] += item
      case _:
        raise Exception("Unknown Item Name!")
  return result

def create_excel(config:dict[Any], cells: list[list[str]], path: Path) -> None:
  """
  Excelデータを出力する

  Parameters
  ----
  config: 設定データを示す構造体
  cells: 試験項目を示すテーブルデータ
  path: 出力パス
  """
  wb = openpyxl.Workbook()
  ws = wb.worksheets[-1]
  noindex  = cells[0].index("No")
  # define styles
  headdesign = styles.PatternFill(patternType='solid', fgColor=config["Headers"]["BackColor"], bgColor=config["Headers"]["BackColor"])
  headfont = styles.Font(color=config["Headers"]["TextColor"])
  side = styles.Side(style="thin", color="000000")
  border = styles.Border(side, side, side, side)
  # insert caption
  font = {}
  for n, v in config["Sheet"].items():
    match n:
      case n if n.startswith("Font"): font[n[4].lower() + n[5:]] = v
      case "Caption":
        ws.cell(1, 1).value = v
      case "Height":
        ws.row_dimensions[1].height = v    
      case "Name":
        ws.title = v
    if font != {}:
      ws.cell(1, 1).font = styles.Font(**font)
  # fill header
  if "TestResult" in config["Headers"]:
    lc = len(config["Headers"]["TestResult"]["Labels"])
    for c in range(config["Headers"]["TestResult"]["PrintCount"]):
      sc = cells[0].index(config["Headers"]["TestResult"]["Labels"][0]) + c * lc + 1
      ec = sc + lc - 1
      cellobj = ws.cell(START_ROW - 1, sc)
      cellobj.value = config["Headers"]["TestResult"]["Title"].format(c+1)
      ws.merge_cells(f"{cellobj.column_letter}{START_ROW - 1}:{ws.cell(START_ROW - 1, ec).column_letter}{START_ROW - 1}")
      cellobj.alignment = styles.Alignment(horizontal="center") 
      cellobj.fill = headdesign
      cellobj.font = headfont
  # output cells
  for r, line in enumerate(cells):
    print(f"> {line[noindex]}")
    for c, cell in enumerate(line):
      cellobj = ws.cell(r + START_ROW, c + 1)
      # extension width
      calcsize = (len(cell if type(cell) is str else max(cell, key=len)) + 2) * 1.4
      dimensions = ws.column_dimensions[cellobj.column_letter]
      if not "\n" in cell and dimensions.width < calcsize:
          dimensions.width = calcsize
      # set text
      if type(cell) is list:
        cell = "\n".join(cell)
      cellobj.value = cell
      align = {
        "vertical": "top",
        "horizontal": "left",
        "wrapText": True
      }
      cellobj.alignment = styles.Alignment(**align) 
      if r == 0:
        # set style
        cellobj.fill = headdesign
        cellobj.font = headfont
      cellobj.border = border
    if len(cells[0]) - len(line) > 0:
      for c in range(len(cells[0]) - len(line)):
        ws.cell(r + START_ROW, c + 1 + len(line)).border = border
  if "ColumnSet" in config:
    adjusttable(ws, config["ColumnSet"])
  if not path.parent.exists(): path.parent.mkdir()
  wb.save(path)

def adjusttable(sheet: worksheet.Worksheet, replace_table: dict[Any]) -> None:
  """
  テーブルの書式を調整する

  Parameters
  ----
  sheet: 調整対象のワークシート
  replace_table: 置換用テーブル
  """
  def applyProperties(conf: dict[str,Any], cell: cell.Cell | None=None) -> tuple[dict[str,str],dict[str,str], str]:
    font = {}
    align = {}
    newvalue = None
    for n, v in conf.items():
      match n:
        case n if n.startswith("Font"): font[n[4].lower() + n[5:]] = v
        case n if n.startswith("Align"): align[n[5].lower() + n[6:]] = v
        case "Replace":
          if cell: 
            cell.value = v 
          else: 
            newvalue = v
        case "Width":
          sheet.column_dimensions[cell.column_letter].width = v
    return (font, align, newvalue)
  print(">> Adjustment")
  for c, col in enumerate(sheet.columns):
    print(f"> {col[START_ROW - 1].value}")
    if col[START_ROW - 1].value in replace_table:
      conf = dictknife.deepmerge(replace_table["Common"], replace_table[col[START_ROW - 1].value])
    else:
      conf = replace_table["Common"]
    if conf != {}:
      if "Header" in conf:
        font, align, _ = applyProperties(conf["Header"], col[START_ROW - 1])
        col[START_ROW - 1].font = styles.Font(**font)
        col[START_ROW - 1].alignment = styles.Alignment(**align)
        if col[START_ROW - 2].value:
          if "TestResultHeader" in replace_table:
            for n, v in replace_table["TestResultHeader"].items():
              match n:
                case "AlignHorizontal": align["horizontal"] = v
                case "AlignVertical": align["vertical"] = v
                case "Height": sheet.row_dimensions[START_ROW - 1].height = v
          col[START_ROW - 2].font = styles.Font(**font)
          col[START_ROW - 2].alignment = styles.Alignment(**align)
      if "Body" in conf:
        cfont, calign, value = applyProperties(conf["Body"])
        font = styles.Font(**cfont)
        align= styles.Alignment(**calign)
        for r in range(sheet.max_row):
          if r < START_ROW: continue
          cellobj = sheet.cell(r + 1, c + 1)
          cellobj.font = font
          cellobj.alignment = align
          if value: cellobj.value = value        

  if "HeaderRow" in replace_table and "Height" in replace_table["HeaderRow"]: sheet.row_dimensions[START_ROW].height = replace_table["HeaderRow"]["Height"]

def expandvars(text: str | list[list[str]], consts: dict[str,str]):
  """
  テーブルないし文字列内の変数データを展開する

  Parameters
  ----
  text: テーブルないし文字列
  consts: 定数を示す辞書データ

  Returns
  ----
  試験項目を示すテーブルデータ。
  """
  if type(text) is list:
    for i, item in enumerate(text):
      text[i] = expandvars(item, consts)
  else:
    for n, v in consts.items():
      text = text.replace("{{" +n + "}}", v)
  return text

if __name__ == "__main__":
  p = ArgumentParser(description="Test Sheet Creation Tool")
  p.add_argument("tests", type=str, help="Markdown file that defines a test item.")
  p.add_argument("-o", "--out", required=True, type=str, help="Excel file output destination.")
  p.add_argument("-c", "--config", default="sample/config.yml", type=str, help="Configured file that defines basic information of the test vote.")
  args = p.parse_args()
  print("> prepare")

  with open(args.config, mode="r", encoding="utf-8") as f: config = yaml.safe_load(f)
  with open(args.tests, mode="r", encoding="utf-8") as f: lines = f.read()

  exams = generate_testlist(lines)
  cells = cells_normalization(config["Headers"]["TestItemsLabel"], exams)
  if "TestResult" in config["Headers"]:
    cells = add_examcells(config["Headers"]["TestResult"], cells)
  if "Rearrange" in config:
    cells = rearrange_cells(config["Headers"], cells, config["Rearrange"])
  if "Consts" in config:
    cells = expandvars(cells, config["Consts"])
  create_excel(config, cells, path=Path(args.out))
  print("> finished!")

