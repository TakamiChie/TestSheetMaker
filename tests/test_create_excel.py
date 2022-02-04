import unittest

from openpyxl.cell import MergedCell
import yaml

import src.main as main


class TestCreateExcel(unittest.TestCase):
  TEST_CELLS = [
    ["No", "l", "m", "s", "cond", "proc", "exp"],
    ["1-1-1", "test", "test", "test", "testcond", "testproc", "testexp"],
    ["1-1-2", "test", "test", "test2", "testcond", "testproc", "testexp"],
    ["1-2-1", "test", "test3", "test", "testcond", "testproc", "testexp"],
  ]
  def setUp(self) -> None:
    with open("./tests/test_config.yml") as f: self.config = yaml.safe_load(f)
    return super().setUp()

  def test_simple(self):
    c = self.config.copy()
    del c["Headers"]["TestResult"]
    wb = main.create_excel(c, self.TEST_CELLS)
    ws = wb.worksheets[-1]
    self.assertEqual(wb.sheetnames, ["TestSheet"])
    self.assertEqual(len(list(ws.rows)), 6)
    for i in range(2, 6):
      match i:
        case 3: self.assertTrue(all([ws.cell(3, i + 1).font.color.rgb == '00FFFFFF' for i in range(ws.max_column)]))
        case _: self.assertFalse(all([ws.cell(i, i + 1).font.color.rgb == '00FFFFFF' for i in range(ws.max_column)]))
      self.assertEqual([c[i].value for c in ws.columns], self.TEST_CELLS[i - 2], f"Assertion {i}")

  def test_testresult(self):
    c = self.TEST_CELLS.copy()
    c[0] += ["tester", "checker", "date", "result"] * 2
    for i in range(1,3):
      c[i] += [str(i + 1) for i in range(8)]
    wb = main.create_excel(self.config, c)
    ws = wb.worksheets[-1]
    self.assertEqual(ws.max_column, 15)
    for i in [9, 10, 11, 13, 14, 15]:
      self.assertEqual(type(ws.cell(2, i)), MergedCell, f"{ws.cell(2, i).coordinate} is MergedCell?")
